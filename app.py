import streamlit as st
import pandas as pd
import requests
import io
import zipfile
import os
import re
from datetime import datetime, timezone, timedelta
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage

# ==============================================================================
# 1. CONFIGURACIÓN DE LA PÁGINA Y PANEL DE ESTILOS CSS
# ==============================================================================
icono_pestana = "sergemLogo.ico" if os.path.exists("sergemLogo.ico") else "sergemLogo.png"
st.set_page_config(page_title="SERGEM - Generador automático de documentos", page_icon=icono_pestana, layout="wide")

st.markdown("""
    <style>
    [data-testid="stHeader"] { display: none !important; }
    [data-testid="stToolbar"] { display: none !important; }
    .stAppDeployButton { display: none !important; }
    #MainMenu { display: none !important; }
    footer { display: none !important; }
    div[class*="viewerBadge"], [data-testid="stAppCreatorBadge"] { display: none !important; }
    
    .stApp { background-color: #F4F6F9; }
    .block-container {
        background-color: #FFFFFF;
        padding: 2rem 3rem;
        border-radius: 12px;
        box-shadow: 0px 4px 15px rgba(0, 0, 0, 0.05);
        margin-top: 2rem;
        margin-bottom: 2rem;
    }
    h1, h2, h3, p, span, label, div { color: #1E293B; }
    
    .stSelectbox > div > div > div {
        background-color: #f8fafc !important;
        border: 1px solid #cbd5e1 !important;
        border-radius: 8px !important;
        color: #1e293b !important;
    }
    
    div[data-testid="stTextInput"] input {
        border: 2px solid #E3000F !important;
        background-color: #fff1f2 !important;
        color: #E3000F !important;
        font-size: 1.15rem !important;
        font-weight: 800 !important;
        padding: 0.75rem !important;
        border-radius: 8px !important;
    }
    
    div.stButton > button {
        border-radius: 8px !important;
        font-weight: 800 !important;
        border: none !important;
        transition: transform 0.1s ease !important;
    }
    div.stButton > button:active { transform: scale(0.98) !important; }
    
    div.stButton > button[kind="primary"] { 
        background-color: #E3000F !important; 
        color: white !important; 
        padding: 0.75rem 1.5rem;
        font-size: 1.1rem;
    }
    div.stButton > button[kind="primary"]:hover { background-color: #B3000C !important; box-shadow: 0px 4px 10px rgba(227, 0, 15, 0.3); }
    
    div.stButton > button[kind="secondary"] { 
        background-color: #94a3b8 !important; 
        color: white !important; 
    }
    div.stButton > button[kind="secondary"]:hover { background-color: #64748b !important; }
    
    .metric-box {
        background-color: #f1f5f9; padding: 15px; border-radius: 8px; text-align: center; border: 1px solid #e2e8f0;
    }
    </style>
    """, unsafe_allow_html=True)

GAS_URL = "https://script.google.com/macros/s/AKfycbyqJtrmVdNT1rxTobg6q_WoJCwMpp40hdIzJeEm4dKNLBgDVxwEY95T0EIoBu_qo8FB/exec"

def obtener_fecha_actual():
    meses = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
    zona_colombia = timezone(timedelta(hours=-5))
    hoy = datetime.now(zona_colombia)
    return f"{hoy.day} DE {meses[hoy.month - 1]} DE {hoy.year}"

@st.cache_data(ttl=600) 
def cargar_datos(url):
    try:
        req_url = f"{url}?t={int(datetime.now().timestamp())}"
        response = requests.get(req_url)
        return response.json()
    except Exception as e:
        return None

def limpiar_dinero(val):
    if pd.isna(val) or val == "": return 0.0
    s = str(val).upper().replace('$', '').replace(',', '').replace('.', '').replace(' ', '')
    try: return float(s)
    except: return 0.0

def limpiar_texto(txt):
    if pd.isna(txt): return ""
    return re.sub(r'\s+', ' ', str(txt).upper().strip())

def obtener_horas(row):
    for col in ['HORAS', 'CANTIDAD DE HORAS', 'CANTIDAD HORAS', 'TOTAL HORAS', 'CANTIDAD', 'NUMERO DE HORAS']:
        if col in row.index:
            try:
                val = float(row[col])
                if not pd.isna(val): return val
            except: pass
    return 1.0

def get_pdf_bytes(pdf_obj):
    out = pdf_obj.output()
    return out.encode('latin-1') if isinstance(out, str) else bytes(out)

# ==============================================================================
# GENERACIÓN DE ARCHIVO PLANO BANCARIO (ESTRUCTURA EXACTA TXT)
# ==============================================================================
def generar_txt_banco(df_banco):
    lineas = []
    zona_colombia = timezone(timedelta(hours=-5))
    hoy = datetime.now(zona_colombia)
    fecha_ymd = hoy.strftime("%Y%m%d")
    fecha_dmy = hoy.strftime("%d/%m/%Y").ljust(13)
    
    num_registros = len(df_banco)
    suma_total = int(df_banco['VALOR_NETO_A_PAGAR'].sum() * 100)
    
    header = (
        f"1000000900561833I               225NOMINA    "
        f"{fecha_ymd}AA{fecha_ymd}"
        f"{num_registros:06d}"
        f"00000000000000000"
        f"{suma_total:017d}"
        f"81016173001D"
    )
    lineas.append(header)
    
    for _, row in df_banco.iterrows():
        cedula = str(row['NIT_BENEFICIARIO']).replace('.', '').replace(' ', '').replace(',', '')
        cedula = cedula.zfill(15)[:15]
        nombre = str(row['NOMBRE_BENEFICIARIO']).upper().ljust(30)[:30]
        banco = "005600078" 
        
        # AQUÍ ESTÁ LA MAGIA: Se eliminan guiones y espacios en blanco del número de cuenta
        cuenta = str(row['NUMERO_CUENTA']).replace("'", "").replace("-", "").replace(" ", "").strip().ljust(18)[:18]
        
        tipo_cta = "37" if "AHORRO" in str(row['TIPO_CUENTA']).upper() else "27"
        valor = int(row['VALOR_NETO_A_PAGAR'] * 100)
        valor_str = f"{valor:017d}"
        filler = "00000                                                                                               000000000000000                           "
        
        linea = (
            f"6{cedula}{nombre}{banco}{cuenta}{tipo_cta}{valor_str}"
            f"{fecha_ymd}NOMINA   {fecha_dmy}{filler}"
        )
        lineas.append(linea)
        
    return "\n".join(lineas)

# ==============================================================================
# LÓGICA DE PDFS 
# ==============================================================================
def agregar_pagina_pdf_cuenta_cobro(pdf, datos):
    pdf.add_page()
    pdf.set_text_color(0, 0, 0)
    
    pdf.set_font("helvetica", "", 11)
    pdf.cell(0, 6, f"{datos['ciudad']} {datos['fecha_emision']}".upper(), 0, 1, 'R')
    pdf.ln(12)
    
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(0, 6, "SERGEM MENSAJERIA S.A.S.", 0, 1, 'C')
    pdf.set_font("helvetica", "", 11)
    pdf.cell(0, 6, "NIT: 900.561.833-1", 0, 1, 'C')
    pdf.ln(8)
    
    pdf.cell(0, 6, "DEBE A:", 0, 1, 'C')
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(0, 6, str(datos['nombre_prestador']).upper(), 0, 1, 'C')
    pdf.set_font("helvetica", "", 11)
    pdf.cell(0, 6, f"C.C / NIT {datos['cedula_prestador']}", 0, 1, 'C')
    pdf.ln(8)

    pdf.set_font("helvetica", "B", 11)
    pdf.cell(80, 6, "VALOR BASE:", 0, 0)
    pdf.set_font("helvetica", "", 11)
    pdf.cell(0, 6, f"$ {datos['ingreso_base']:,.0f}", 0, 1)

    if datos.get('fuera_perimetro', 0) > 0:
        pdf.set_font("helvetica", "B", 11)
        pdf.cell(80, 6, "FUERA DE PERÍMETRO:", 0, 0)
        pdf.set_font("helvetica", "", 11)
        pdf.cell(0, 6, f"$ {datos['fuera_perimetro']:,.0f}", 0, 1)

    if datos.get('retefuente', 0) > 0:
        pdf.set_font("helvetica", "B", 11)
        pdf.cell(80, 6, "MENOS RETEFUENTE (1%):", 0, 0)
        pdf.set_font("helvetica", "", 11)
        pdf.set_text_color(227, 0, 15)
        pdf.cell(0, 6, f"$ -{datos['retefuente']:,.0f}", 0, 1)
        pdf.set_text_color(0, 0, 0)

    if datos.get('ica', 0) > 0:
        pdf.set_font("helvetica", "B", 11)
        pdf.cell(80, 6, "MENOS RETEICA (1%):", 0, 0)
        pdf.set_font("helvetica", "", 11)
        pdf.set_text_color(227, 0, 15)
        pdf.cell(0, 6, f"$ -{datos['ica']:,.0f}", 0, 1)
        pdf.set_text_color(0, 0, 0)

    pdf.ln(2)
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(80, 8, "NETO SERVICIOS:", 0, 0)
    pdf.cell(0, 8, f"$ {datos['neto_pagar']:,.0f}", 0, 1)
    pdf.ln(6)
    
    pdf.set_font("helvetica", "B", 11)
    pdf.cell(25, 6, "CONCEPTO:", 0, 1)
    pdf.set_font("helvetica", "", 10)
    pdf.multi_cell(0, 5, f"SERVICIO DE MENSAJERÍA PRESTADO EN EL CORTE DE {datos['corte_fechas']}, DETALLADO A CONTINUACIÓN:")
    pdf.ln(2)
    
    pdf.set_fill_color(227, 0, 15)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("helvetica", "B", 10)
    pdf.cell(85, 6, "CONDUCTOR / DETALLE", 1, 0, 'C', fill=True)
    pdf.cell(35, 6, "CÉDULA", 1, 0, 'C', fill=True)
    pdf.cell(25, 6, "CANTIDAD", 1, 0, 'C', fill=True)
    pdf.cell(45, 6, "VALOR NETO", 1, 1, 'C', fill=True)

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("helvetica", "", 9)
    for c in datos['conductores']:
        pdf.cell(85, 6, c['nombre_conductor'][:40], 1, 0, 'L')
        pdf.cell(35, 6, c['cedula_conductor'], 1, 0, 'C')
        pdf.cell(25, 6, f"{c['horas']:g} Horas", 1, 0, 'C')
        pdf.cell(45, 6, f"$ {c['neto_pagar']:,.0f}", 1, 1, 'R')
        
    for fpu in datos.get('fpu_items', []):
        pdf.cell(85, 6, f"F. Perímetro: {fpu['destino'][:25]}", 1, 0, 'L')
        pdf.cell(35, 6, "N/A", 1, 0, 'C')
        pdf.cell(25, 6, f"{fpu['cantidad']:g} Viaje(s)", 1, 0, 'C')
        pdf.cell(45, 6, f"$ {fpu['neto']:,.0f}", 1, 1, 'R')

    pdf.ln(8)
    
    if datos.get('anticipos', 0) > 0 or datos.get('otros_descuentos', 0) > 0:
        if datos.get('anticipos', 0) > 0:
            pdf.set_font("helvetica", "B", 11)
            pdf.cell(80, 6, "MENOS ANTICIPOS:", 0, 0)
            pdf.set_font("helvetica", "", 11)
            pdf.set_text_color(227, 0, 15)
            pdf.cell(0, 6, f"$ -{datos['anticipos']:,.0f}", 0, 1)
            pdf.set_text_color(0, 0, 0)
            
        if datos.get('otros_descuentos', 0) > 0:
            pdf.set_font("helvetica", "B", 11)
            pdf.cell(80, 6, "MENOS OTROS DESCUENTOS (ARL):", 0, 0)
            pdf.set_font("helvetica", "", 11)
            pdf.set_text_color(227, 0, 15)
            pdf.cell(0, 6, f"$ -{datos['otros_descuentos']:,.0f}", 0, 1)
            pdf.set_text_color(0, 0, 0)

        pdf.ln(2)
        pdf.set_font("helvetica", "B", 12)
        pdf.set_text_color(227, 0, 15)
        pdf.cell(80, 8, "TOTAL A CONSIGNAR:", 0, 0)
        pdf.cell(0, 8, f"$ {datos['neto_final']:,.0f}", 0, 1)
        pdf.set_text_color(0, 0, 0)
        pdf.ln(6)
    
    pdf.set_font("helvetica", "", 11)
    pdf.cell(0, 6, "Autorizo me sea consignado en:", 0, 1)
    pdf.set_font("helvetica", "B", 11)
    pdf.cell(0, 6, f"CUENTA # {datos['num_cuenta']} - {datos['tipo_cuenta'].upper()}", 0, 1)
    pdf.cell(0, 6, f"BANCO: {datos['banco'].upper()}", 0, 1)
    pdf.cell(0, 6, f"TITULAR: {datos['nombre_titular_banco']} (C.C/NIT: {datos['cedula_titular_banco']})", 0, 1)
    pdf.ln(15)
    
    pdf.set_font("helvetica", "", 11)
    pdf.cell(0, 6, "Atentamente,", 0, 1)
    pdf.ln(10)
    pdf.set_font("helvetica", "B", 11)
    pdf.cell(80, 5, str(datos['nombre_prestador']).upper(), "T", 1, "L")
    pdf.set_font("helvetica", "", 11)
    pdf.cell(80, 5, f"C.C / NIT {datos['cedula_prestador']}", 0, 1, "L")

def agregar_pagina_pdf_doc_equivalente(pdf, datos):
    pdf.add_page()
    
    try:
        if os.path.exists('sergemLogo.png'):
            pdf.image('sergemLogo.png', 10, 8, w=45)
    except: pass

    pdf.set_font('helvetica', 'B', 10)
    pdf.cell(0, 5, "DOCUMENTO EQUIVALENTE A LA FACTURA DE VENTA", 0, 1, 'R')
    pdf.set_font('helvetica', '', 9)
    pdf.cell(0, 4, "(DECRETO 522 DE 2003)", 0, 1, 'R')
    pdf.cell(0, 4, "DOCUMENTO SOPORTE EN ADQUISICIONES A NO OBLIGADOS A FACTURAR", 0, 1, 'R')
    pdf.ln(2)
    
    pdf.set_font('helvetica', 'B', 11)
    pdf.set_text_color(227, 0, 15)
    pdf.cell(0, 6, f"CONSECUTIVO NO: {datos['id']}", 0, 1, 'R')
    pdf.set_text_color(0, 0, 0)
    pdf.ln(5)

    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(35, 5, "Fecha de Expedición:")
    pdf.set_font('helvetica', '', 9)
    pdf.cell(0, 5, datos['fecha_emision'], 0, 1)
    pdf.ln(3)

    pdf.set_fill_color(51, 51, 51)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(0, 6, " INFORMACIÓN DE LA EMPRESA (COMPRADOR)", 1, 1, 'L', fill=True)
    pdf.set_text_color(0, 0, 0)
    
    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(25, 6, "Razón Social:", 1)
    pdf.set_font('helvetica', '', 9)
    pdf.cell(100, 6, "SERGEM MENSAJERIA S.A.S.", 1)
    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(15, 6, "NIT:", 1)
    pdf.set_font('helvetica', '', 9)
    pdf.cell(0, 6, "900.561.833-1", 1, 1)

    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(25, 6, "Dirección:", 1)
    pdf.set_font('helvetica', '', 9)
    pdf.cell(60, 6, "CRA 62 9 235", 1)
    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(20, 6, "Teléfono:", 1)
    pdf.set_font('helvetica', '', 9)
    pdf.cell(45, 6, "3994620", 1)
    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(15, 6, "Ciudad:", 1)
    pdf.set_font('helvetica', '', 9)
    pdf.cell(0, 6, "CALI", 1, 1)
    pdf.ln(4)

    pdf.set_fill_color(51, 51, 51)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(0, 6, " DATOS DEL BENEFICIARIO / PROVEEDOR (VENDEDOR)", 1, 1, 'L', fill=True)
    pdf.set_text_color(0, 0, 0)

    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(25, 6, "Nombre:", 1)
    pdf.set_font('helvetica', '', 9)
    pdf.cell(100, 6, datos['nombre_prestador'][:45], 1)
    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(20, 6, "C.C / NIT:", 1)
    pdf.set_font('helvetica', '', 9)
    pdf.cell(0, 6, datos['cedula_prestador'], 1, 1)
    
    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(25, 6, "Ciudad:", 1)
    pdf.set_font('helvetica', '', 9)
    pdf.cell(100, 6, datos['ciudad'], 1)
    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(20, 6, "Conductores:", 1)
    pdf.set_font('helvetica', '', 9)
    
    nombres_conds = ", ".join([c['nombre_conductor'] for c in datos['conductores']])
    if len(nombres_conds) > 25: nombres_conds = nombres_conds[:22] + "..."
    pdf.cell(0, 6, nombres_conds, 1, 1)
    pdf.ln(6)

    # Detalle de Items
    pdf.set_fill_color(227, 0, 15)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(10, 6, "Ítem", 1, 0, 'C', fill=True)
    pdf.cell(90, 6, "Concepto", 1, 0, 'C', fill=True)
    pdf.cell(20, 6, "Cantidad", 1, 0, 'C', fill=True)
    pdf.cell(35, 6, "V. Unitario", 1, 0, 'C', fill=True)
    pdf.cell(0, 6, "V. Total", 1, 1, 'C', fill=True)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('helvetica', '', 9)

    item_idx = 1
    for c in datos['conductores']:
        v_unitario = c['ingreso_base'] / c['horas'] if c['horas'] > 0 else c['ingreso_base']
        pdf.cell(10, 6, str(item_idx), 1, 0, 'C')
        pdf.cell(90, 6, f"Servicio mensajería - {c['nombre_conductor'][:25]}", 1, 0, 'L')
        pdf.cell(20, 6, f"{c['horas']:g}", 1, 0, 'C')
        pdf.cell(35, 6, f"$ {v_unitario:,.0f}", 1, 0, 'R')
        pdf.cell(0, 6, f"$ {c['ingreso_base']:,.0f}", 1, 1, 'R')
        item_idx += 1

    for fpu in datos.get('fpu_items', []):
        pdf.cell(10, 6, str(item_idx), 1, 0, 'C')
        pdf.cell(90, 6, f"Fuera Perímetro: {fpu['destino'][:20]}", 1, 0, 'L')
        pdf.cell(20, 6, f"{fpu['cantidad']:g}", 1, 0, 'C')
        pdf.cell(35, 6, f"$ {fpu['valor_unitario']:,.0f}", 1, 0, 'R')
        pdf.cell(0, 6, f"$ {fpu['total']:,.0f}", 1, 1, 'R')
        item_idx += 1

    pdf.ln(2)
    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(120, 6, "", 0, 0)
    pdf.cell(35, 6, "SUBTOTAL:", 1, 0, 'R')
    pdf.cell(0, 6, f"$ {datos['ingreso_bruto_total']:,.0f}", 1, 1, 'R')

    pdf.cell(120, 6, "", 0, 0)
    pdf.cell(35, 6, "IVA (19%):", 1, 0, 'R')
    pdf.cell(0, 6, "$ 0", 1, 1, 'R')

    pdf.cell(120, 6, "", 0, 0)
    pdf.cell(35, 6, "RETEIVA:", 1, 0, 'R')
    pdf.cell(0, 6, "$ 0", 1, 1, 'R')

    pdf.cell(120, 6, "", 0, 0)
    pdf.cell(35, 6, "RTE FTE (1%):", 1, 0, 'R')
    val_rte = -datos['retefuente'] if datos['retefuente'] > 0 else 0
    pdf.cell(0, 6, f"$ {val_rte:,.0f}", 1, 1, 'R')

    pdf.cell(120, 6, "", 0, 0)
    pdf.cell(35, 6, "RETEICA (1%):", 1, 0, 'R')
    val_ica = -datos['ica'] if datos['ica'] > 0 else 0
    pdf.cell(0, 6, f"$ {val_ica:,.0f}", 1, 1, 'R')

    pdf.set_fill_color(244, 246, 249)
    pdf.cell(120, 6, "", 0, 0)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(35, 6, "NETO SERVICIOS:", 1, 0, 'R', fill=True)
    pdf.cell(0, 6, f"$ {datos['neto_pagar']:,.0f}", 1, 1, 'R', fill=True)

    if datos.get('anticipos', 0) > 0:
        pdf.cell(120, 6, "", 0, 0)
        pdf.cell(35, 6, "ANTICIPOS:", 1, 0, 'R')
        pdf.set_text_color(227, 0, 15)
        pdf.cell(0, 6, f"$ -{datos['anticipos']:,.0f}", 1, 1, 'R')
        pdf.set_text_color(0, 0, 0)
        
    if datos.get('otros_descuentos', 0) > 0:
        pdf.cell(120, 6, "", 0, 0)
        pdf.cell(35, 6, "OTROS DESC (ARL):", 1, 0, 'R')
        pdf.set_text_color(227, 0, 15)
        pdf.cell(0, 6, f"$ -{datos['otros_descuentos']:,.0f}", 1, 1, 'R')
        pdf.set_text_color(0, 0, 0)
        
    if datos.get('anticipos', 0) > 0 or datos.get('otros_descuentos', 0) > 0:
        pdf.cell(120, 8, "", 0, 0)
        pdf.set_text_color(227, 0, 15)
        pdf.cell(35, 8, "TOTAL A CONSIGNAR:", 1, 0, 'R', fill=True)
        pdf.set_font('helvetica', 'B', 11)
        pdf.cell(0, 8, f"$ {datos['neto_final']:,.0f}", 1, 1, 'R', fill=True)
        pdf.set_text_color(0, 0, 0)

    pdf.ln(12)
    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(80, 5, "________________________________________________", 0, 1)
    pdf.cell(80, 5, "FIRMA PRESTADOR DEL SERVICIO", 0, 1)
    pdf.cell(80, 5, f"C.C. / NIT: {datos['cedula_prestador']}", 0, 1)
    pdf.cell(80, 5, f"NOMBRE: {datos['nombre_prestador']}", 0, 1)

def construir_hoja_documento_equivalente_excel(ws, datos):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="E3000F", end_color="E3000F", fill_type="solid")
    dark_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    bold_font = Font(bold=True)
    border_thin = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'), top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    try:
        if os.path.exists('sergemLogo.png'):
            img = XLImage('sergemLogo.png')
            img.width = 150
            img.height = 60
            ws.add_image(img, 'B2')
    except: pass

    ws.merge_cells('D2:H4')
    ws['D2'] = "DOCUMENTO EQUIVALENTE A LA FACTURA DE VENTA\n(DECRETO 522 DE 2003)\nDOCUMENTO SOPORTE EN ADQUISICIONES A NO OBLIGADOS A FACTURAR"
    ws['D2'].font = Font(bold=True, size=11, color="1E293B")
    ws['D2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    hoy = datetime.now(timezone(timedelta(hours=-5)))
    ws['B6'] = "Fecha de Expedición:"; ws['B6'].font = bold_font
    ws['C6'] = "Año:"; ws['D6'] = hoy.year; ws['E6'] = "Mes:"; ws['F6'] = f"{hoy.month:02d}"; ws['G6'] = "Día:"; ws['H6'] = f"{hoy.day:02d}"
    
    ws['G5'] = "CONSECUTIVO NO:"; ws['G5'].font = Font(bold=True, size=11, color="E3000F"); ws['G5'].alignment = right_align
    ws['H5'] = datos['id']; ws['H5'].font = Font(bold=True, size=12, color="E3000F"); ws['H5'].alignment = center_align

    ws['B9'] = " INFORMACIÓN DE LA EMPRESA (COMPRADOR)"
    ws['B9'].font = header_font; ws['B9'].fill = dark_fill
    ws.merge_cells('B9:H9')
    ws['B10'] = "Razón Social:"; ws['B10'].font = bold_font; ws['C10'] = "SERGEM MENSAJERIA S.A.S."; ws.merge_cells('C10:E10')
    ws['G10'] = "NIT:"; ws['G10'].font = bold_font; ws['H10'] = "900.561.833-1"

    ws['B13'] = " DATOS DEL BENEFICIARIO / PROVEEDOR (VENDEDOR)"
    ws['B13'].font = header_font; ws['B13'].fill = dark_fill; ws.merge_cells('B13:H13')

    ws['B14'] = "Nombre:"; ws['B14'].font = bold_font; ws['C14'] = datos['nombre_prestador']; ws.merge_cells('C14:E14')
    ws['G14'] = "C.C / NIT:"; ws['G14'].font = bold_font; ws['H14'] = datos['cedula_prestador']
    
    ws['G16'] = "Conductores:"; ws['G16'].font = bold_font
    nombres_conds = ", ".join([c['nombre_conductor'] for c in datos['conductores']])
    if len(nombres_conds) > 25: nombres_conds = nombres_conds[:22] + "..."
    ws['H16'] = nombres_conds

    fila = 18
    for i, h in enumerate(["Ítem", "Concepto", "Cantidad", "V. Unitario", "V. Total"]):
        c = ['B', 'C', 'F', 'G', 'H'][i] + str(fila)
        ws[c] = h; ws[c].font = header_font; ws[c].fill = header_fill; ws[c].alignment = center_align; ws[c].border = border_thin
    ws.merge_cells(f'C{fila}:E{fila}')
    
    fila += 1
    item_idx = 1
    for c in datos['conductores']:
        ws[f'B{fila}'] = item_idx
        ws[f'C{fila}'] = f"Servicio mensajería - {c['nombre_conductor']}"
        ws.merge_cells(f'C{fila}:E{fila}')
        ws[f'F{fila}'] = float(c['horas']) 
        v_unitario = c['ingreso_base'] / c['horas'] if c['horas'] > 0 else c['ingreso_base']
        ws[f'G{fila}'] = v_unitario
        ws[f'H{fila}'] = c['ingreso_base']
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']: ws[f'{col}{fila}'].border = border_thin
        ws[f'G{fila}'].number_format = '"$"#,##0'; ws[f'H{fila}'].number_format = '"$"#,##0'
        ws[f'B{fila}'].alignment = center_align; ws[f'C{fila}'].alignment = left_align; ws[f'F{fila}'].alignment = center_align
        fila += 1
        item_idx += 1

    for fpu in datos.get('fpu_items', []):
        ws[f'B{fila}'] = item_idx
        ws[f'C{fila}'] = f"Fuera Perímetro: {fpu['destino']}"
        ws.merge_cells(f'C{fila}:E{fila}')
        ws[f'F{fila}'] = fpu['cantidad']
        ws[f'G{fila}'] = fpu['valor_unitario']
        ws[f'H{fila}'] = fpu['total']
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']: ws[f'{col}{fila}'].border = border_thin
        ws[f'G{fila}'].number_format = '"$"#,##0'; ws[f'H{fila}'].number_format = '"$"#,##0'
        ws[f'B{fila}'].alignment = center_align; ws[f'C{fila}'].alignment = left_align; ws[f'F{fila}'].alignment = center_align
        fila += 1
        item_idx += 1
        
    fila += 1
    
    totales = [
        ("SUBTOTAL:", datos['ingreso_bruto_total']), 
        ("IVA (19%):", ""), ("RETEIVA:", ""), 
        ("RTE FTE (1%):", -datos['retefuente'] if datos['retefuente']>0 else 0),
        ("RETEICA (1%):", -datos['ica'] if datos['ica']>0 else 0), 
        ("NETO SERVICIOS:", datos['neto_pagar'])
    ]
    
    if datos.get('anticipos', 0) > 0: totales.append(("MENOS ANTICIPOS:", -datos['anticipos']))
    if datos.get('otros_descuentos', 0) > 0: totales.append(("OTROS DESC (ARL):", -datos['otros_descuentos']))
    
    if datos.get('anticipos', 0) > 0 or datos.get('otros_descuentos', 0) > 0:
        totales.append(("TOTAL A CONSIGNAR:", datos['neto_final']))
    
    fila_firma = fila + len(totales) + 2
    for label, valor in totales:
        ws[f'G{fila}'] = label; ws[f'H{fila}'] = valor
        ws[f'G{fila}'].font = bold_font; ws[f'G{fila}'].alignment = right_align
        ws[f'G{fila}'].border = border_thin; ws[f'H{fila}'].border = border_thin
        if valor != "": ws[f'H{fila}'].number_format = '"$"#,##0'
        
        es_resaltado = (label == "TOTAL A CONSIGNAR:") or (label == "NETO SERVICIOS:" and datos.get('anticipos', 0) == 0 and datos.get('otros_descuentos', 0) == 0)
        if es_resaltado:
            ws[f'G{fila}'].font = Font(bold=True, color="E3000F"); ws[f'H{fila}'].font = Font(bold=True, size=12)
            ws[f'H{fila}'].fill = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")
        fila += 1

    ws[f'B{fila_firma}'] = "________________________________________________"
    ws[f'B{fila_firma+1}'] = "FIRMA PRESTADOR DEL SERVICIO"
    ws[f'B{fila_firma+1}'].font = bold_font
    ws[f'B{fila_firma+2}'] = f"C.C. / NIT: {datos['cedula_prestador']}"
    ws[f'B{fila_firma+3}'] = f"NOMBRE: {datos['nombre_prestador']}"

    ws.column_dimensions['B'].width = 16; ws.column_dimensions['C'].width = 12; ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12; ws.column_dimensions['F'].width = 10; ws.column_dimensions['G'].width = 22; ws.column_dimensions['H'].width = 22

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT; ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True; ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.5; ws.page_margins.right = 0.5; ws.page_margins.top = 0.5; ws.page_margins.bottom = 0.5

# ==============================================================================
# PROCESO MATEMÁTICO PRINCIPAL (AGRUPADO POR TITULAR)
# ==============================================================================
def obtener_nombre_columna(df, opciones):
    for op in opciones:
        for col in df.columns:
            if limpiar_texto(col) == limpiar_texto(op):
                return col
    return None

def calcular_valores_agrupados(grupo_df, df_fuera, corte_seleccionado, col_prestador, col_ced_prestador, col_tit_banco, col_ced_banco, col_estado, col_anticipos, col_otros_desc):
    conductores = []
    fpu_items_doc = []
    conductores_procesados_fpu = set()
    
    es_nuevo = False
    suma_neto = 0
    suma_bruto = 0
    suma_fuera_bruto = 0
    suma_retefuente = 0
    suma_ica = 0
    suma_horas = 0
    
    suma_anticipos = 0
    suma_otros_desc = 0

    row_titular = grupo_df.iloc[0]
    
    nombre_prestador = str(row_titular.get(col_prestador, 'S/N')).strip()
    cedula_prestador = str(row_titular.get(col_ced_prestador, '')).strip()
    
    nombre_titular_banco = str(row_titular.get(col_tit_banco, nombre_prestador)).strip()
    cedula_titular_banco = str(row_titular.get(col_ced_banco, cedula_prestador)).strip()
    
    banco = str(row_titular.get('BANCO', '')).strip()
    tipo_cuenta = str(row_titular.get('TIPO CUENTA', '')).strip()
    num_cuenta = str(row_titular.get('NO. CUENTA', '')).strip()
    ciudad_titular = str(row_titular.get('CIUDAD', '')).upper().strip()

    for _, row in grupo_df.iterrows():
        if col_estado and str(row.get(col_estado, '')).strip().upper() == 'NUEVO':
            es_nuevo = True
            
        if col_anticipos: suma_anticipos += limpiar_dinero(row.get(col_anticipos, 0))
        if col_otros_desc: suma_otros_desc += limpiar_dinero(row.get(col_otros_desc, 0))

        ingreso_neto_esperado = limpiar_dinero(row.get('TOTAL A PAGAR', 0))
        nombre_conductor = str(row.get('CONDUCTOR', '')).strip().upper()

        ciudad = str(row.get('CIUDAD', '')).upper().strip()
        cedula_conductor = str(row.get('CÉDULA', row.get('CEDULA', ''))).strip()
        horas = obtener_horas(row)

        porcentaje_retefuente = 0.01
        porcentaje_ica = 0.01 if ciudad == 'CALI' else 0.0
        tasa_total_impuestos = porcentaje_retefuente + porcentaje_ica

        fuera_perimetro_neto = 0.0
        if "MILTON" in nombre_conductor and not df_fuera.empty and nombre_conductor not in conductores_procesados_fpu:
            conductores_procesados_fpu.add(nombre_conductor)
            
            col_dest = obtener_nombre_columna(df_fuera, ['FUERA PERIMETRO CEDI', 'DESTINO', 'CIUDAD', 'LUGAR'])
            col_val = obtener_nombre_columna(df_fuera, ['VALOR', 'PRECIO'])
            col_cant = obtener_nombre_columna(df_fuera, ['CANTIDAD', 'CANT'])
            
            if col_dest and col_val and col_cant:
                for _, f_row in df_fuera.iterrows():
                    try:
                        cant = float(f_row.get(col_cant, 0))
                        if cant > 0:
                            destino = str(f_row.get(col_dest, ''))
                            valor_uni_neto = limpiar_dinero(f_row.get(col_val, 0))
                            tot_neto = cant * valor_uni_neto
                            
                            tot_bruto = round(tot_neto / (1 - tasa_total_impuestos))
                            val_uni_bruto = round(valor_uni_neto / (1 - tasa_total_impuestos))
                            
                            fuera_perimetro_neto += tot_neto
                            suma_fuera_bruto += tot_bruto
                            
                            fpu_items_doc.append({
                                'destino': destino,
                                'cantidad': cant,
                                'valor_unitario': val_uni_bruto,
                                'total': tot_bruto,
                                'neto': tot_neto
                            })
                    except: pass

        total_neto_esperado = ingreso_neto_esperado + fuera_perimetro_neto

        ingreso_bruto_total = round(total_neto_esperado / (1 - tasa_total_impuestos)) if total_neto_esperado > 0 else 0
        retefuente = round(ingreso_bruto_total * porcentaje_retefuente)
        ica = round(ingreso_bruto_total * porcentaje_ica)

        fpu_bruto_cond = round(fuera_perimetro_neto / (1 - tasa_total_impuestos)) if fuera_perimetro_neto > 0 else 0.0
        ingreso_base_bruta = ingreso_bruto_total - fpu_bruto_cond
        
        neto_total_conductor = ingreso_bruto_total - retefuente - ica
        neto_solo_horas = neto_total_conductor - fuera_perimetro_neto

        if ingreso_neto_esperado > 0 or nombre_conductor != "":
            conductores.append({
                'nombre_conductor': nombre_conductor if nombre_conductor else nombre_prestador,
                'cedula_conductor': cedula_conductor if cedula_conductor else cedula_prestador,
                'horas': horas,
                'ingreso_base': ingreso_base_bruta,
                'neto_pagar': neto_solo_horas
            })

        suma_neto += neto_total_conductor
        suma_bruto += ingreso_bruto_total
        suma_retefuente += retefuente
        suma_ica += ica
        suma_horas += horas

    # Garantizar que nadie con $0 (como Jilmer) sea descartado por error de fila vacía
    if not conductores: 
        conductores.append({
            'nombre_conductor': nombre_prestador,
            'cedula_conductor': cedula_prestador,
            'horas': suma_horas if suma_horas > 0 else 1.0,
            'ingreso_base': 0,
            'neto_pagar': 0
        })

    neto_final = suma_neto - suma_anticipos - suma_otros_desc

    return {
        'es_nuevo': es_nuevo,
        'anticipos': suma_anticipos,
        'otros_descuentos': suma_otros_desc,
        'neto_final': neto_final,
        'nombre_prestador': nombre_prestador,
        'cedula_prestador': cedula_prestador,
        'nombre_titular_banco': nombre_titular_banco,
        'cedula_titular_banco': cedula_titular_banco,
        'banco': banco,
        'tipo_cuenta': tipo_cuenta,
        'num_cuenta': num_cuenta,
        'ciudad': ciudad_titular,
        'conductores': conductores,
        'fpu_items': fpu_items_doc,
        'ingreso_base': suma_bruto - suma_fuera_bruto,
        'fuera_perimetro': suma_fuera_bruto,
        'ingreso_bruto_total': suma_bruto,
        'retefuente': suma_retefuente,
        'ica': suma_ica,
        'neto_pagar': suma_neto,
        'total_horas': suma_horas
    }

# ==============================================================================
# INTERFAZ DE USUARIO
# ==============================================================================
col1, col2 = st.columns([1, 4])
with col1:
    try:
        if os.path.exists("sergemLogo.png"): st.image("sergemLogo.png", use_column_width=True)
        elif os.path.exists("sergemLogo_2.png"): st.image("sergemLogo_2.png", use_column_width=True)
    except: pass
with col2:
    st.title("Generador Automático de Documentos")
    st.markdown("**SERGEM Mensajería S.A.S.**")

if st.button("🔄 Sincronizar Base de Datos", key="btn_sync", type="secondary"):
    cargar_datos.clear() 
    st.cache_data.clear()
    st.rerun()

data_cruda = cargar_datos(GAS_URL)
if not data_cruda:
    st.error("Error conectando a Google Sheets. Revise la URL o los permisos del Apps Script.")
    st.stop()

# ==============================================================================
# SEPARACIÓN ESTRICTA DE BASES DE DATOS
# ==============================================================================
df_pagos_completo = pd.DataFrame(data_cruda.get('pagos', []))
df_bd_maestra = pd.DataFrame(data_cruda.get('bd', []))
df_fuera = pd.DataFrame(data_cruda.get('fueras_perimetro', []))

if df_pagos_completo.empty:
    st.warning("No se encontraron datos en la pestaña PAGOS PERSONAL POR SERVICIOS.")
    st.stop()

df_pagos_completo.columns = df_pagos_completo.columns.str.strip().str.upper()
if not df_bd_maestra.empty: df_bd_maestra.columns = df_bd_maestra.columns.str.strip().str.upper()
if not df_fuera.empty: df_fuera.columns = df_fuera.columns.str.strip().str.upper()

if 'CORTE' in df_pagos_completo.columns:
    df_pagos_completo['CORTE'] = df_pagos_completo['CORTE'].astype(str).str.strip().str.upper()

col_prestador = obtener_nombre_columna(df_pagos_completo, ['A NOMBRE DE QUIEN HACE CUENTA DE COBRO', 'NOMBRE PRESTADOR', 'A NOMBRE DE QUIEN HACE CUENTA'])
col_cedula_prestador = obtener_nombre_columna(df_pagos_completo, ['CÉDULA DE CUENTA DE COBRO', 'CEDULA DE CUENTA DE COBRO'])
col_titular_banco = obtener_nombre_columna(df_pagos_completo, ['NOMBRE TITULAR CUENTA BANCARIA', 'NOMBRE_TITULAR'])
col_cedula_banco = obtener_nombre_columna(df_pagos_completo, ['CÉDULA TITULAR', 'CEDULA TITULAR'])

col_estado = obtener_nombre_columna(df_pagos_completo, ['ESTADO', 'ESTADO_EMPLEADO'])
col_anticipos = obtener_nombre_columna(df_pagos_completo, ['ANTICIPOS', 'ANTICIPO'])
col_otros_desc = obtener_nombre_columna(df_pagos_completo, ['OTROS DESCUENTOS', 'OTROS_DESCUENTOS', 'DESCUENTOS'])

if not col_prestador or not col_titular_banco:
    st.error("Faltan las columnas que diferencian a quien cobra del titular del banco. Verifique sus nombres en el Sheets.")
    st.stop()

cortes_disponibles = [c for c in df_pagos_completo['CORTE'].unique() if str(c).strip() != "" and str(c).lower() != "nan"]

st.divider()
modo_trabajo = st.radio("⚙️ Modo de trabajo:", 
                        ["🗂️ Generación Masiva (Paquete Gerencial)", 
                         "👤 Vista Previa Individual",
                         "⏱️ Actualizador de Horas Automático (Excel a Drive)"], horizontal=True)

# ==============================================================================
# MODO ACTUALIZADOR DE HORAS (100% DINÁMICO A LA ESTRUCTURA DE LA BD)
# ==============================================================================
if "Actualizador" in modo_trabajo:
    st.markdown("### ⏱️ Depurador y Actualizador de Horas (De Sistema a Drive)")
    st.info("Sube el archivo Excel biométrico. El programa construirá una tabla depurada basándose en tu pestaña BD para que la pegues en PAGOS PERSONAL POR SERVICIOS.")
    
    nuevo_corte = st.text_input("✍️ Escriba el nombre exacto del Corte a generar (Ej: 1 AL 15 AGOSTO):")
    archivo_horas = st.file_uploader("📥 Sube el reporte de horas en formato Excel (.xlsx)", type=["xlsx", "xls"])
    
    if archivo_horas and nuevo_corte:
        try:
            df_raw = pd.read_excel(archivo_horas)
            
            col_cc = obtener_nombre_columna(df_raw, ['CC', 'CEDULA', 'CÉDULA'])
            col_horas = obtener_nombre_columna(df_raw, ['TOTAL_HORAS', 'TOTAL HORAS', 'HORAS'])
            
            if col_cc and col_horas:
                df_raw[col_cc] = pd.to_numeric(df_raw[col_cc], errors='coerce')
                df_raw = df_raw.dropna(subset=[col_cc])
                
                agg_dict = {col: 'first' for col in df_raw.columns if col != col_cc and col != col_horas}
                agg_dict[col_horas] = 'sum'
                grouped = df_raw.groupby(col_cc, as_index=False).agg(agg_dict)
                
                columnas_destino = [c for c in df_bd_maestra.columns if str(c).strip() != "" and "UNNAMED" not in str(c).upper()]
                
                col_ced_bd = obtener_nombre_columna(df_bd_maestra, ['CÉDULA', 'CEDULA', 'C.C.', 'C.C', 'CC'])
                col_horas_bd = obtener_nombre_columna(df_bd_maestra, ['NÚMERO DE HORAS', 'NUMERO DE HORAS', 'HORAS', 'TOTAL HORAS'])
                col_corte_bd = obtener_nombre_columna(df_bd_maestra, ['CORTE', 'PERIODO'])
                col_total_bd = obtener_nombre_columna(df_bd_maestra, ['TOTAL A PAGAR', 'TOTAL_A_PAGAR'])
                col_val_hora_bd = obtener_nombre_columna(df_bd_maestra, ['VALOR HORA', 'VALOR_HORA'])
                
                result_rows = []
                for _, row in grouped.iterrows():
                    cc = row[col_cc]
                    horas = row[col_horas]
                    
                    match = pd.DataFrame()
                    if col_ced_bd:
                        ced_bd = df_bd_maestra[col_ced_bd].astype(str).str.replace(".0", "", regex=False).str.strip()
                        ced_match = str(cc).replace(".0", "").strip()
                        match = df_bd_maestra[ced_bd == ced_match]
                    
                    new_row = {}
                    
                    for col in columnas_destino:
                        val_final = ""
                        alias_busqueda = [col]
                        
                        if col == "CLIENTE": alias_busqueda.extend(["EMPRESA", "PUNTO_VENTA"])
                        if col == "CONDUCTOR": alias_busqueda.extend(["MENSAJERO", "NOMBRE"])
                        if col == "VALOR HORA": alias_busqueda.extend(["VALOR_HORA"])
                        if col == "ESTADO": alias_busqueda.extend(["ESTADO_EMPLEADO"])
                        
                        col_raw_match = obtener_nombre_columna(df_raw, alias_busqueda)
                        if col_raw_match and pd.notna(row[col_raw_match]) and str(row[col_raw_match]).strip() != "":
                            val_final = row[col_raw_match]
                        elif not match.empty:
                            bd_row = match.iloc[0]
                            bd_col_match = obtener_nombre_columna(df_bd_maestra, alias_busqueda)
                            if bd_col_match and pd.notna(bd_row[bd_col_match]) and str(bd_row[bd_col_match]).strip() != "":
                                val_final = bd_row[bd_col_match]
                                
                        new_row[col] = val_final
                    
                    if col_ced_bd: new_row[col_ced_bd] = int(cc) if cc else ""
                    if col_horas_bd: new_row[col_horas_bd] = round(horas, 2)
                    if col_corte_bd: new_row[col_corte_bd] = nuevo_corte.strip().upper()
                    
                    val_hora = 0
                    if col_val_hora_bd:
                        val_hora = limpiar_dinero(new_row.get(col_val_hora_bd, 0))
                        
                    if col_total_bd:
                        new_row[col_total_bd] = round(horas * val_hora, 0) if val_hora > 0 else 0
                        
                    result_rows.append(new_row)
                    
                df_res = pd.DataFrame(result_rows)
                
                st.success(f"✅ ¡Cruce Exitoso 100% Dinámico! Se extrajeron las horas y cruzaron con todas las columnas actuales de BD.")
                
                excel_out = io.BytesIO()
                df_res.to_excel(excel_out, index=False, sheet_name="PAGOS PERSONAL POR SERVICIOS")
                excel_out.seek(0)
                
                st.download_button(
                    label="📥 DESCARGAR BASE DEPURADA (EXCEL PARA DRIVE)",
                    data=excel_out,
                    file_name=f"Base_Depurada_SERGEM_{nuevo_corte.replace(' ', '_').replace('/', '-')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )
                
            else:
                st.error("El archivo subido no tiene la estructura biométrica. Faltan las columnas 'CC' o 'TOTAL_HORAS'.")
        except Exception as e:
            st.error(f"Error procesando el archivo: {e}")

# ==============================================================================
# MODOS DE GENERACIÓN DE DOCUMENTOS (INDIVIDUAL Y LOTE)
# ==============================================================================
else:
    if not cortes_disponibles:
        st.stop()
    
    corte_seleccionado = st.selectbox("📅 Seleccione el Corte a procesar:", cortes_disponibles)
    df_pagos_corte = df_pagos_completo[df_pagos_completo['CORTE'] == corte_seleccionado]

    if df_pagos_corte['CONDUCTOR'].str.contains('MILTON', na=False).any():
        st.info("💡 **Aviso Importante:** Se detectó a Milton Javier Cortes. Asegúrate de tener una columna llamada **CANTIDAD** en tu pestaña *FUERAS PERIMETRO /ADIC* indicando cuántos viajes hizo a cada destino para que salgan desglosados en su cuenta.")

    # ------------------------------------------------------------------------------
    # MODO INDIVIDUAL
    # ------------------------------------------------------------------------------
    if "Individual" in modo_trabajo:
        temp_df = df_pagos_corte[[col_prestador, col_cedula_prestador, col_titular_banco, col_cedula_banco]].copy()
        temp_df = temp_df.fillna("SIN DATO") 
        titulares_unicos = temp_df.drop_duplicates(subset=[col_cedula_banco])
        
        lista_opciones = []
        for _, row in titulares_unicos.iterrows():
            lbl = f"{row[col_prestador]} (C.C: {row[col_cedula_prestador]}) 🏦 Pago a cuenta de CC: {row[col_cedula_banco]}"
            lista_opciones.append(lbl)
            
        opcion_seleccionada = st.selectbox("Busque o seleccione el prestador (titular de la cuenta de cobro):", sorted(lista_opciones))
        
        if st.button("🔍 Calcular y Previsualizar"):
            ced_banco_seleccionada = opcion_seleccionada.split("CC: ")[1].strip()
            
            grupo_titular = df_pagos_corte[df_pagos_corte[col_cedula_banco].astype(str).str.replace(".0", "", regex=False).str.strip() == ced_banco_seleccionada]
            
            calculos = calcular_valores_agrupados(grupo_titular, df_fuera, corte_seleccionado, col_prestador, col_cedula_prestador, col_titular_banco, col_cedula_banco, col_estado, col_anticipos, col_otros_desc)
            
            if not calculos:
                st.warning("Este titular tiene saldo neto en $0 para este corte sin anticipos reportados.")
            else:
                st.markdown(f"### Resumen Financiero: {calculos['nombre_prestador']}")
                if calculos['es_nuevo']:
                    st.info("👤 **ESTADO NUEVO DETECTADO:** Esta persona se incluirá en el paquete exclusivo de nuevos.")
                
                cA, cB, cC, cD = st.columns(4)
                cA.markdown(f"<div class='metric-box'><b>BASE BRUTA</b><br>${calculos['ingreso_base']:,.0f}</div>", unsafe_allow_html=True)
                cB.markdown(f"<div class='metric-box'><b>DESCUENTOS LEGALES</b><br>${calculos['retefuente']+calculos['ica']:,.0f}</div>", unsafe_allow_html=True)
                cC.markdown(f"<div class='metric-box'><b>ANTICIPOS/ARL</b><br>${calculos['anticipos']+calculos['otros_descuentos']:,.0f}</div>", unsafe_allow_html=True)
                cD.markdown(f"<div class='metric-box' style='background-color:#E3000F; color:white;'><b>TOTAL CONSIGNAR</b><br>${calculos['neto_final']:,.0f}</div>", unsafe_allow_html=True)
                
                if calculos['fuera_perimetro'] > 0:
                    st.success(f"🚚 **¡Valor detectado!** Se sumaron **${calculos['fuera_perimetro']:,.0f}** adicionales detallados correspondientes a Salidas Fuera de Perímetro.")
                
                if calculos['neto_final'] <= 0:
                    st.warning("⚠️ El total a consignar da **$0 o negativo** debido a los descuentos. No se enviará este pago al archivo del banco, pero sí puedes descargar los documentos.")

            datos_doc = calculos.copy()
            datos_doc.update({'id': "PREVIEW", 'fecha_emision': obtener_fecha_actual(), 'corte_fechas': corte_seleccionado})
            
            pdf_ct = FPDF(); agregar_pagina_pdf_cuenta_cobro(pdf_ct, datos_doc)
            pdf_eq = FPDF(); agregar_pagina_pdf_doc_equivalente(pdf_eq, datos_doc)
            
            colBtn1, colBtn2 = st.columns(2)
            colBtn1.download_button("📥 Descargar Cuenta de Cobro (PDF)", data=get_pdf_bytes(pdf_ct), file_name=f"Cuenta_{calculos['cedula_prestador']}.pdf", mime="application/pdf", use_container_width=True)
            colBtn2.download_button("📥 Descargar Doc. Equivalente (PDF)", data=get_pdf_bytes(pdf_eq), file_name=f"DocEq_{calculos['cedula_prestador']}.pdf", mime="application/pdf", use_container_width=True)

    # ------------------------------------------------------------------------------
    # MODO MASIVO (LOTE)
    # ------------------------------------------------------------------------------
    elif "Masiva" in modo_trabajo:
        if st.button("🚀 Procesar Lote General", use_container_width=True, type="primary"):
            mensaje_carga = st.info(f"📥 Procesando la información y empaquetando archivos de forma inteligente...")
            
            try:
                pagos_procesados_banco = []
                nuevos_detectados = []
                ceros_detectados = []
                
                ignorados = count_banco = count_nuevos = count_ceros = 0
                fecha_actual = obtener_fecha_actual() 
                
                # SEPARACIÓN DE INSTANCIAS PARA GENERACIÓN TOTALMENTE INDEPENDIENTE
                pdf_ct_banco = FPDF(); pdf_eq_banco = FPDF()
                wb_eq_banco = openpyxl.Workbook(); wb_eq_banco.remove(wb_eq_banco.active)
                
                pdf_ct_nuevos = FPDF(); pdf_eq_nuevos = FPDF()
                wb_eq_nuevos = openpyxl.Workbook(); wb_eq_nuevos.remove(wb_eq_nuevos.active)
                
                pdf_ct_ceros = FPDF(); pdf_eq_ceros = FPDF()
                wb_eq_ceros = openpyxl.Workbook(); wb_eq_ceros.remove(wb_eq_ceros.active)
                
                cedulas_banco_unicas = df_pagos_corte[col_cedula_banco].dropna().unique()
                contador = 1
                
                for ced_banco in cedulas_banco_unicas:
                    grupo_titular = df_pagos_corte[df_pagos_corte[col_cedula_banco] == ced_banco]
                    
                    calculos = calcular_valores_agrupados(grupo_titular, df_fuera, corte_seleccionado, col_prestador, col_cedula_prestador, col_titular_banco, col_cedula_banco, col_estado, col_anticipos, col_otros_desc)
                    
                    if not calculos:
                        ignorados += 1
                        continue
                    
                    datos_doc = calculos.copy()
                    datos_doc.update({'id': str(contador).zfill(3), 'fecha_emision': fecha_actual, 'corte_fechas': corte_seleccionado})
                    nombre_pestana = f"{contador}_{datos_doc['nombre_prestador'][:20]}".replace(":", "").replace("/", "-")

                    # 1. SI ES NUEVO: Se compila en su paquete exclusivo de Nuevos (PDFs + Excel)
                    if datos_doc['es_nuevo']:
                        agregar_pagina_pdf_cuenta_cobro(pdf_ct_nuevos, datos_doc)
                        agregar_pagina_pdf_doc_equivalente(pdf_eq_nuevos, datos_doc)
                        ws = wb_eq_nuevos.create_sheet(title=nombre_pestana)
                        construir_hoja_documento_equivalente_excel(ws, datos_doc)
                        nuevos_detectados.append(datos_doc)
                        count_nuevos += 1

                    # 2. CLASIFICACIÓN FINANCIERA (SALDO CERO VS BANCO)
                    if datos_doc['neto_final'] <= 0:
                        agregar_pagina_pdf_cuenta_cobro(pdf_ct_ceros, datos_doc)
                        agregar_pagina_pdf_doc_equivalente(pdf_eq_ceros, datos_doc)
                        ws = wb_eq_ceros.create_sheet(title=nombre_pestana)
                        construir_hoja_documento_equivalente_excel(ws, datos_doc)
                        ceros_detectados.append(datos_doc)
                        count_ceros += 1
                        
                    else:
                        # ENTRA AL ARCHIVO Y PLANO BANCARIO (Incluye nuevos con saldo positivo para dar exactamente 63)
                        agregar_pagina_pdf_cuenta_cobro(pdf_ct_banco, datos_doc)
                        agregar_pagina_pdf_doc_equivalente(pdf_eq_banco, datos_doc)
                        ws = wb_eq_banco.create_sheet(title=nombre_pestana)
                        construir_hoja_documento_equivalente_excel(ws, datos_doc)
                        
                        pagos_procesados_banco.append({
                            'NIT_BENEFICIARIO': datos_doc['cedula_titular_banco'],
                            'NOMBRE_BENEFICIARIO': datos_doc['nombre_titular_banco'],
                            'BANCO_DESTINO': datos_doc['banco'],
                            'TIPO_CUENTA': datos_doc['tipo_cuenta'],
                            'NUMERO_CUENTA': datos_doc['num_cuenta'],
                            'VALOR_NETO_A_PAGAR': datos_doc['neto_final'],
                            'FECHA_PAGO': datetime.now(timezone(timedelta(hours=-5))).strftime("%Y/%m/%d"),
                            'CONCEPTO': 'NOMINA'
                        })
                        count_banco += 1
                        
                    contador += 1
                
                # PREPARACIÓN DE DESCARGAS PRINCIPALES (BANCO)
                zip_cuentas_banco_io = io.BytesIO()
                if count_banco > 0:
                    with zipfile.ZipFile(zip_cuentas_banco_io, "w", zipfile.ZIP_DEFLATED) as zipf:
                        zipf.writestr("Cuentas_de_Cobro_Aprobadas.pdf", get_pdf_bytes(pdf_ct_banco))
                zip_cuentas_banco_io.seek(0)
                
                zip_eq_banco_io = io.BytesIO()
                if count_banco > 0:
                    with zipfile.ZipFile(zip_eq_banco_io, "w", zipfile.ZIP_DEFLATED) as zipf:
                        zipf.writestr("Documentos_Equivalentes_Aprobados.pdf", get_pdf_bytes(pdf_eq_banco))
                        excel_io = io.BytesIO(); wb_eq_banco.save(excel_io); excel_io.seek(0)
                        zipf.writestr("Documentos_Equivalentes_Excel.xlsx", excel_io.read())
                zip_eq_banco_io.seek(0)
                
                df_banco = pd.DataFrame(pagos_procesados_banco)
                texto_banco = generar_txt_banco(df_banco) if len(df_banco) > 0 else "SIN REGISTROS VALIDOS PARA EL BANCO"
                
                mensaje_carga.empty() 
                st.success(f"✅ ¡Éxito! Procesamiento finalizado. **{count_banco}** pagos aprobados listos para pago en banco.")
                st.divider()

                # BLOQUE RESTAURADO PARA DOÑA YESENIA: LOS 3 BOTONES DE NUEVOS
                if count_nuevos > 0:
                    df_nuevos = pd.DataFrame([{
                        'NOMBRES Y APELLIDOS': d['nombre_prestador'], 
                        'CÉDULA': d['cedula_prestador'], 
                        'BANCO': d['banco'], 
                        'TIPO CUENTA': d['tipo_cuenta'], 
                        'NO. CUENTA': d['num_cuenta']
                    } for d in nuevos_detectados])
                    
                    excel_nuevos_io = io.BytesIO()
                    df_nuevos.to_excel(excel_nuevos_io, index=False, sheet_name="PERSONAL NUEVO")
                    excel_nuevos_io.seek(0)

                    st.error(f"🚨 **ATENCIÓN - SE DETECTARON {count_nuevos} PERSONAS NUEVAS**\n\nSus soportes contables se generaron de forma independiente y también se compiló su listado para Don José. Utiliza estos botones para descargar su información:")
                    
                    colN1, colN2, colN3 = st.columns(3)
                    colN1.download_button("1️⃣ 📥 Listado Excel (Para Don José)", data=excel_nuevos_io, file_name="Listado_Nuevos.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                    colN2.download_button("2️⃣ 📥 Cuentas de Cobro (Solo Nuevos)", data=get_pdf_bytes(pdf_ct_nuevos), file_name="Cuentas_Cobro_Nuevos.pdf", mime="application/pdf", use_container_width=True)
                    
                    zip_eq_nuevos_io = io.BytesIO()
                    with zipfile.ZipFile(zip_eq_nuevos_io, "w", zipfile.ZIP_DEFLATED) as zipf:
                        zipf.writestr("Docs_Equivalentes_Nuevos.pdf", get_pdf_bytes(pdf_eq_nuevos))
                        excel_io = io.BytesIO(); wb_eq_nuevos.save(excel_io); excel_io.seek(0)
                        zipf.writestr("Docs_Equivalentes_Nuevos_Excel.xlsx", excel_io.read())
                    zip_eq_nuevos_io.seek(0)
                    colN3.download_button("3️⃣ 📥 Docs Equivalentes (Solo Nuevos)", data=zip_eq_nuevos_io, file_name="Docs_Equivalentes_Nuevos.zip", mime="application/zip", use_container_width=True)
                    st.divider()
                
                # BLOQUE PARA DOÑA YESENIA: BOTONES EXCLUSIVOS PARA SALDOS EN CERO
                if count_ceros > 0:
                    nombres_ceros = "\n* ".join([f"👤 {d['nombre_prestador']} (C.C: {d['cedula_prestador']})" for d in ceros_detectados])
                    st.warning(f"⚠️ **SE DETECTARON {count_ceros} SALDOS EN CERO O NEGATIVOS**\n\nPersonas con deducciones totales por anticipos o ARL. Fueron excluidos del archivo plano, pero puedes descargar sus soportes aquí para contabilizar el cruce de los anticipos:\n* {nombres_ceros}")
                    
                    colC1, colC2 = st.columns(2)
                    colC1.download_button("1️⃣ 📥 Cuentas de Cobro (Saldos Cero)", data=get_pdf_bytes(pdf_ct_ceros), file_name="Cuentas_Cobro_Ceros.pdf", mime="application/pdf", use_container_width=True)
                    
                    zip_eq_ceros_io = io.BytesIO()
                    with zipfile.ZipFile(zip_eq_ceros_io, "w", zipfile.ZIP_DEFLATED) as zipf:
                        zipf.writestr("Docs_Equivalentes_Ceros.pdf", get_pdf_bytes(pdf_eq_ceros))
                        excel_io = io.BytesIO(); wb_eq_ceros.save(excel_io); excel_io.seek(0)
                        zipf.writestr("Docs_Equivalentes_Ceros_Excel.xlsx", excel_io.read())
                    zip_eq_ceros_io.seek(0)
                    colC2.download_button("2️⃣ 📥 Docs Equivalentes (Saldos Cero)", data=zip_eq_ceros_io, file_name="Docs_Equivalentes_Ceros.zip", mime="application/zip", use_container_width=True)
                    st.divider()

                # DESCARGAS GENERALES (PERSONAL APROBADO)
                st.markdown(f"### 📥 Soportes Contables Aprobados (Los {count_banco} del Banco)")
                colD1, colD2, colD3 = st.columns(3)
                
                colD1.download_button(
                    label="1️⃣ Soportes: Cuentas de Cobro (.ZIP)",
                    data=zip_cuentas_banco_io,
                    file_name=f"Cuentas_Cobro_Aprobadas_{corte_seleccionado.replace(' ', '_').replace('/', '-')}.zip",
                    mime="application/zip",
                    use_container_width=True,
                    disabled=(count_banco == 0)
                )
                
                colD2.download_button(
                    label="2️⃣ Soportes: Docs. Equivalentes (.ZIP)",
                    data=zip_eq_banco_io,
                    file_name=f"Docs_Equivalentes_Aprobados_{corte_seleccionado.replace(' ', '_').replace('/', '-')}.zip",
                    mime="application/zip",
                    use_container_width=True,
                    disabled=(count_banco == 0)
                )
                
                colD3.download_button(
                    label="3️⃣ Archivo Plano Banco (.TXT)",
                    data=texto_banco,
                    file_name=f"Plano_Banco_{corte_seleccionado.replace(' ', '_').replace('/', '-')}.txt",
                    mime="text/plain",
                    use_container_width=True,
                    disabled=(count_banco == 0)
                )

            except Exception as e:
                mensaje_carga.empty()
                st.error(f"Error en el proceso: {e}")
